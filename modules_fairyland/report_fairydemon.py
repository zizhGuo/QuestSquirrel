import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
import functools

import sys
print(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'transform')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'transform')))

from transform.base_transform import BaseTransform
'''
'''
def adjust_column_width(ws):
    # 计算并调整每列宽度
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                # 计算当前单元格内容的长度
                cell_length = len(str(cell.value))
                # 如果是新列，或者当前单元格内容长度超过已记录的最大长度，则更新
                if cell.column_letter not in column_widths or cell_length > column_widths[cell.column_letter]:
                    column_widths[cell.column_letter] = cell_length
    # 设置列宽（略大于最大长度以避免内容截断）
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width*2  # 可根据需要调整额外宽度

class ReportSchedular:
    def __init__(self, config, root_path, module) -> None:
        self.report_generator = []
        for sub_report, _sub_config in config[module.report_module].items():
            _ = ReportGenerator(config, root_path, module, sub_report)
            self.report_generator.append(_)
    def run(self):
        for i in range(len(self.report_generator)):
            print('run report generator')
            try:
                print(f'generating report: {i}')
                self.report_generator[i].run()
                print('-----------------------------------')
                print('-----------------------------------')
                print('-----------------------------------')
                print('-----------------------------------')
                print('-----------------------------------')
                print('-----------------------------------')
            except Exception as e:
                print(f'gen report run failed. {i}')
                print(e)


class ReportGenerator:
    def __init__(self, config, root_path, module, sub_report):
        self.module = module
        self.sub_report = sub_report
        if config[module.report_module][sub_report].get('transform') is not None:
            print('successfully import transform class.')
            self.transform_dict = config[module.report_module][sub_report]['transform']
            print('transform_dict: {}'.format(self.transform_dict))
        self.tables = config[module.report_module][sub_report]['source_tables']
        # filter out the 'na' tables
        # self.tables = [table for table in self.tables if table != 'na']
        self.num_tables = len(self.tables)
        self.output_dir =  config[module.report_module][sub_report]['output_file']
        self.worksheet =  config[module.report_module][sub_report]['worksheet']
        self.ws2title =  config[module.report_module][sub_report]['ws2title']
        self.column =  config[module.report_module][sub_report]['column']
        self.col2names =  config[module.report_module][sub_report]['col2names']
        self.root_path = root_path
        self.temp_save_path = config[module.report_module][sub_report]['temp_save_path']
        self.end_dt = config['end_dt']
 
    def run(self):
        wb = Workbook()
        wb.remove(wb.active)  # 移除初始的空worksheet
        ws_data = {}
        for table, ws in zip(self.tables, self.worksheet):
            if ws not in ws_data:
                ws_data[ws] = []
            ws_data[ws].append(table)
        print(f"ws_data: {ws_data}")
        # for ws, files in sorted(ws_data.items(), key = lambda x : int(x[0][2:])):
        for ws, files in ws_data.items():
            ws = wb.create_sheet(title=self.ws2title[ws])
            start_row = 1
            for file in files:
                _dir = os.path.join(self.root_path, self.temp_save_path, self.end_dt, file)
                if not os.path.exists(_dir):
                    print(f"file {file} not existed")
                    continue
                df = pd.read_excel(_dir)

                i = self.tables.index(file)
                # rename columns
                df.columns = self.col2names[self.column[i]]
                
                # insert dataframe edit class
                # df = new_class(df).edit()
                # if dict has tranform class, then create child class and edit
                if hasattr(self, 'transform_dict') and type(self.transform_dict.get(file)) == list:
                    transform_instance = self._create_transform_instance(file)()
                    start_row = transform_instance.run(ws, df, start_row)
                # else use default add_dataframe_to_worksheet
                else:
                    transform_instance = BaseTransform()
                    start_row = transform_instance.run(ws, df, start_row)

        dir_path = os.path.join(self.root_path, self.temp_save_path, self.end_dt)
        file_path = os.path.join(dir_path, self.output_dir)
        print(f"file_path: {file_path}")
        wb.save(f"{file_path}")
    
    def generate_god_batch(self, df_write, sheet_name, mapper, write_xlsx):
        tables = {}
        for sheet_df in zip(sheet_name, df_write):
            if sheet_df[0] == 'na':
                continue
            if not tables.get(sheet_df[0]):
                tables[sheet_df[0]] = [sheet_df[1]]
            else:
                tables[sheet_df[0]].append(sheet_df[1])
        assert mapper.get('files2sheets') is not None, 'files2sheets is not existed'
        for file, sheet_list in mapper['files2sheets'].items():
            tables_new = {}
            for sheet in sheet_list:
                tables_new[sheet] = tables[sheet]
            write_xlsx(tables_new, self.root_path, self.output_dir, file)
    
    def _create_transform_instance(self, file):
        import importlib
        print(self.transform_dict[file][0])
        module = importlib.import_module(self.transform_dict[file][0])
        Class = getattr(module, self.transform_dict[file][1])
        return Class


