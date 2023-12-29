import os
import sys
CURRENT_FILE_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(CURRENT_FILE_DIR)
sys.path.insert(0, CURRENT_FILE_DIR)
sys.path.insert(1, PARENT_DIR)

from modules.arguments import parser
from modules.config import ConfigManager
from modules.schedular_god import TaskScheduler

CONFIG_FILE = 'config_god.yaml'

from datetime import datetime
now = datetime.now()

def main():
    print("entered main god")

    config_path = os.path.join(CURRENT_FILE_DIR, CONFIG_FILE)
    config_manager = ConfigManager(config_path)
    config = config_manager.config

    args = parser.parse_args()
    from datetime import timedelta, datetime
    from dateutil.relativedelta import relativedelta
    def date2str(date, timestamp_format=False):
        if timestamp_format:
            if date.day < 10:
                return f"{date.year}-{date.month}-0{date.day}"
            return f"{date.year}-{date.month}-{date.day}" 
        else:
            if date.day < 10:
                return f"{date.year}{date.month}0{date.day}"
            return f"{date.year}{date.month}{date.day}"
    end_dt = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d"))
    # end_date = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d"), timestamp_format = True)
    # start_dt = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d") - timedelta(days=7))
    # start_date = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d") - timedelta(days=7), timestamp_format = True)
    # pre_start_dt = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d") - timedelta(days=7) - relativedelta(days=365))
    # pre_start_date = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d") - timedelta(days=7) - relativedelta(days=365), timestamp_format = True)
    # # pre_start_dt = date2str(datetime.strptime(args.end_dt, "%Y-%m-%d") - timedelta(days=377))
    # # print("end_dt = {}; start_dt = {}; pre_start_dt = {}".format(end_dt, start_dt, pre_start_dt))
    # config['query']['params']['start_dt'] = start_dt
    # config['query']['params']['start_date'] = start_date
    # config['query']['params']['end_dt'] = end_dt
    # config['query']['params']['end_date'] = end_date
    # config['query']['params']['pre_start_dt'] = pre_start_dt
    # config['query']['params']['pre_start_date'] = pre_start_date


    scheduler = TaskScheduler(config = config, root_path = CURRENT_FILE_DIR, date = end_dt)
    scheduler.run_task()
    scheduler.gen_report(write_xlsx)
    # scheduler.gen_report_test(write_xlsx)


    # scheduler.send_email()
    
    print("main god end")
    pass


def write_xlsx(tables, current_dir, output_dir, file):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Create an Excel workbook
    wb = Workbook()

    # Function to append DataFrames to an Excel sheet at a specific position
    def append_df_to_excel(ws, df, startrow=None, startcol=None):
        rows = dataframe_to_rows(df, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx + startrow, column=c_idx + startcol, value=value)

    # Iterate over the tables dictionary and write each DataFrame to the corresponding sheet
    for sheet_name, dfs in tables.items():
        # Add a new sheet to the workbook with the name sheet_name
        # If sheet_name is 'sheet1' which is the default, use the active sheet
        if sheet_name == 'sheet1':
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # Starting positions for the first DataFrame
        startrow = 0
        startcol = 0
        
        for df in dfs:
            append_df_to_excel(ws, df, startrow=startrow, startcol=startcol)
            # Update startrow to be below the current DataFrame including an offset (e.g., 1 row)
            startrow += len(df.index) + 1

    # Save the workbook   
    wb.save(os.path.join(current_dir, output_dir)+'/'+ file)

def test():
    import pandas as pd
    # Create some example DataFrames
    df1 = pd.DataFrame({
        'A': [1, 2],
        'B': [3, 4]
    })

    df2 = pd.DataFrame({
        'X': [1, 2, 9],
        'Y': [3, 4, 12],
        'Z': [7, 7, 7]
    })  

    df3 = pd.DataFrame({
        'C': [5, 6],
        'D': [7, 8],
        'E': [9, 10]
    })
    
    tables = {
        'sheet1': [df1,df2]
        ,'sheet2': [df3]
    }
    write_xlsx(tables)


   


if __name__ == "__main__":
    main()
    # test()


