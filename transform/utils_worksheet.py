
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import functools

import sys
print(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'transform')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'transform')))

THICK_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'),
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

def df2ws(func):
    """模块装饰器: df转ws装饰器, 用于将df转换为ws并写入ws中
        标准化：粗体标题，边框
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, df, start_row, *args, **kwargs):
        current_row = start_row 
        for _, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(r)
            if current_row == start_row:
                for cell in ws[current_row]:
                    cell.font = Font(bold=True)
                    cell.border = THICK_BORDER
            if 'regular' in kwargs:
                if current_row != start_row:
                    for cell in ws[current_row]:
                        cell.border = THICK_BORDER
            current_row += 1
        ws.append([])

        func(self, ws, df, start_row, *args, **kwargs)
        return current_row + 1
    return wrapper

def border_cell_column(func):
    """模块装饰器: df转ws装饰器, 用于将df转换为ws并写入ws中
        标准化：粗体标题，边框
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, start_row, *args, **kwargs):
        start_row, end_row, column_index= func(self, kwargs.get('bold_section', None), *args, **kwargs)
        if start_row == -1 or end_row == -1 or column_index == -1:
            # quit('bold_section not in kwargs')
            raise IndexError('func: border_cell_column: bold_section not in kwargs') 
        for row in range(start_row, end_row):
            cell = ws[row][0]
            cell.font = Font(bold=True)
                
    return wrapper


def insert_row(func):
    """模块装饰器: 插入行装饰器
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, start_row, content, *args, **kwargs):
        ws.append([content])
        # for cell in ws[start_row]:
        #     cell.font = Font(bold=True)
        func(self, *args, **kwargs)
        return start_row + 1
    return wrapper

def merge_cells(func):
    """装饰器：合并单元格
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, df, start_row, dt, level, target_column, *args, **kwargs):
        cur_row = start_row 
        result = df.groupby([dt, level]).apply(lambda x: pd.Series([x.index[0], x.index[-1]], index=['First_Index', 'Last_Index'])).reset_index()
        result_set = [(row['First_Index']+cur_row+1, row['Last_Index']+cur_row+1) for _, row in result.iterrows()]
        column_index = df.columns.get_loc(target_column)+1
        for start, end in result_set:
            ws.merge_cells(start_row=start, 
                        end_row=end, 
                        start_column=column_index, 
                        end_column=column_index)
        func(self, *args, **kwargs)
    return wrapper

def merge_cells_general(func):
    """装饰器：合并单元格
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, df, start_row, *args, **kwargs):
        # print('groupby_target: {}'.format(groupby_target))
        cur_row = start_row
        if not 'groupby_target' in kwargs:
            print('groupby_target not in kwargs')
            print('kwargs: {}'.format(kwargs))
        else:
            for pair in kwargs['groupby_target']:
                groupby_cols, target_col = pair[0], pair[1]
                result = df.groupby(groupby_cols).apply(lambda x: pd.Series([x.index[0], x.index[-1]], index=['First_Index', 'Last_Index'])).reset_index()
                print('results: {}'.format(result))
                result_set = [(row['First_Index']+cur_row+1, row['Last_Index']+cur_row+1) for _, row in result.iterrows()]
                print('result_set: {}'.format(result_set))
                column_index = df.columns.get_loc(target_col)+1
                print('cur_row: {}'.format(cur_row))
                for start, end in result_set:
                    ws.merge_cells(start_row=start, 
                                end_row=end, 
                                start_column=column_index, 
                                end_column=column_index)
                    if 'alignment' in kwargs:
                        _horizontal, _vertical = kwargs['alignment'].split(',')[0], kwargs['alignment'].split(',')[1]
                        _cell = ws.cell(row=start, column=column_index)
                        _cell.alignment = Alignment(horizontal=_horizontal, 
                                                vertical=_vertical)
        # func(self, ws, df, start_row, groupby_target, *args, **kwargs)
        func(self, ws, df, start_row, *args, **kwargs)
    return wrapper

def color_cells(func):
    """装饰器：单元格上色 (未完成), 返回row index(未完成)
        参数：行索引，列索引，或坐标集合，（df中的索引）
        类型：
            行索引 -> List[int]
            列索引 -> List[int]
            坐标集合 -> List[tuple]
        注意：方法会本地转换索引至worksheet索引
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, df, start_row, *args, **kwargs):
        cur_row = start_row
        if 'row_set_fill' in kwargs:
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            row_set = kwargs['row_set_fill']['row_set']
            assert type(row_set) == list, 'row_set is not a list, exit' 
            for _row_index in row_set:
                _row_index = cur_row + _row_index + 1
                
                # unfinished task: get row_index from kwargs
                # select row given _row_index
                # r = ws[_row_index]
                # r = ws.iloc[_row_index]
                # print(r)
                # print(f"第 {_row_index} 行的内容: {[cell.value for cell in r]}")
                
                for cell in ws[_row_index]:
                    cell.fill = red_fill
        if 'column_set' in kwargs:
            pass
        if 'coord_set' in kwargs:
            pass
        func(self, ws, df, start_row, *args, **kwargs)
        return cur_row # ?????
    return wrapper