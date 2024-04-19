
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
import functools

import sys
print(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'transform')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'transform')))

# current the default style, later extend with the style class
THICK_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'),
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

def arg_test_outer(func):
    """
    """
    @functools.wraps(func)
    def wrapper(self, ws, df, start_row, *args, **kwargs):
        print('enter outer')  
        func(self, ws, df, start_row, *args, **kwargs)
        return 1
    return wrapper


def df2ws(func):
    """模块装饰器: df转ws装饰器, 用于将df转换为ws并写入ws中
        标准化：粗体标题，边框
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, df, start_row, *args, **kwargs):
        current_row = start_row 
        for _, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(r)
            if current_row == start_row:
                for cell in ws[current_row]:
                    cell.font = Font(bold=True)
                    cell.border = THICK_BORDER
            current_row += 1
        ws.append([])

        func(self, ws, df, start_row, *args, **kwargs)
        return current_row + 1
    return wrapper

def insert_row(func):
    """模块装饰器: 插入行装饰器
        先于func执行
    """
    @functools.wraps(func)
    def wrapper(self, ws, start_row, content, *args):
        ws.append([content])
        for cell in ws[start_row]:
            cell.font = Font(bold=True)
        func(self, *args)
        return start_row + 1
    return wrapper

class BaseTransform:
    def __init__(self):
        pass

    @staticmethod
    def arg_test_inner(func):
        """
        """
        @functools.wraps(func)
        def wrapper(self, ws, df, start_row, groupby_target, *args, **kwargs):
            print('enter inner')  
            # print('groupby_target: {}'.format(groupby_target))
            func(self, ws, df, start_row, groupby_target, *args, **kwargs)
        return wrapper

    @staticmethod
    def drop_columns(func):
        """类装饰器：删除指定列
            先于func执行
        """
        @functools.wraps(func)
        def wrapper(self, df, columns, *args):
            print(df)
            df = df.drop(columns, axis=1)
            func(self, *args)
            # 改动
            # df = func(self, df, *args)
            return df
        return wrapper

    @staticmethod
    def select_columns(func):
        """类装饰器: 选择指定列
            先于func执行
        """
        @functools.wraps(func)
        def wrapper(self, df, columns_dict, *args):
            for col, val in columns_dict.items():
                df = df[df[col] == val].reset_index(drop=True)
            func(self, df, columns_dict, *args)
            # df = func(self, df, columns_dict, *args)
            return df
        return wrapper

    @staticmethod
    def merge_cells(func):
        """类装饰器：合并单元格
            先于func执行
        """
        @functools.wraps(func)
        def wrapper(self, ws, df, start_row, dt, level, target_column, *args, **kwargs):
            cur_row = start_row 
            result = df.groupby([dt, level]).apply(lambda x: pd.Series([x.index[0], x.index[-1]], index=['First_Index', 'Last_Index'])).reset_index()
            result_set = [(row['First_Index']+cur_row+1, row['Last_Index']+cur_row+1) for _, row in result.iterrows()]
            column_index = df.columns.get_loc(target_column)+1
            for start, end in result_set:
                ws.merge_cells(start_row=start, 
                            end_row=end, 
                            start_column=column_index, 
                            end_column=column_index)
            func(self, *args, **kwargs)
        return wrapper
    
    @staticmethod
    def merge_cells_general(func):
        """类装饰器：合并单元格
            先于func执行
        """
        @functools.wraps(func)
        def wrapper(self, ws, df, start_row, groupby_target, *args, **kwargs):
            # print('groupby_target: {}'.format(groupby_target))
            cur_row = start_row 
            for pair in groupby_target:
                groupby_cols, target_col = pair[0], pair[1]
                result = df.groupby(groupby_cols).apply(lambda x: pd.Series([x.index[0], x.index[-1]], index=['First_Index', 'Last_Index'])).reset_index()
                print('results: {}'.format(result))
                result_set = [(row['First_Index']+cur_row+1, row['Last_Index']+cur_row+1) for _, row in result.iterrows()]
                print('result_set: {}'.format(result_set))
                column_index = df.columns.get_loc(target_col)+1
                print('cur_row: {}'.format(cur_row))
                for start, end in result_set:
                    ws.merge_cells(start_row=start, 
                                end_row=end, 
                                start_column=column_index, 
                                end_column=column_index)
            # func(self, ws, df, start_row, groupby_target, *args, **kwargs)
            func(self, *args, **kwargs)
        return wrapper

    @df2ws
    def run(self, *args):
        """默认worksheet标准化处理函数, report中实例化后调用"""
        pass