from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
import pandas as pd

# from base_transform import BaseTransform
# from base_transform import df2ws, arg_test_outer, insert_row

from utils_dateframe import drop_columns, select_columns, modify_cell_matched
from utils_worksheet import df2ws, merge_cells_general, insert_row

import logging

class TypeDoubleDictDescriptor:
    """用于描述dict of dict的kwargs类型
    """
    def __init__(self, name, required_keys = None):
        self.name = name
        self.required_keys = required_keys or []

    def __set__(self, instance, value):
        if not isinstance(value, dict):
            raise TypeError(f'{self.name} descriptor set value Failure. kwargs is not a dict: {value}')
        for k, subdict in value.items():
            if not isinstance(subdict, dict):
                raise TypeError(f'{self.name} descriptor set value Failure. subdict is not a dict: {subdict}')
        # set attribute to instance?
        instance.__dict__ = value

    def __get__(self, instance, owner):
        return instance.__dict__

class LevelFailureTransform:
    """境界失败表
    steps:
    *1. delete '[]' cell value
    2. df2ws
    3. merge_cells_general
    """

    def __init__(self, *args, **kwargs):
        self.groupby_target = [
            (['日期', '境界等级'], '总玩家名单'),
            (['日期', '境界等级'], '仅购买了仙魔礼包'),
            (['日期', '境界等级'], '购买了仙魔+常规礼包'),
            (['日期', '境界等级'], '仅购买了常规礼包'),
            (['日期', '境界等级'], '无任何共买记录')
    ]
        self.modify_dict = {
            '总玩家名单': '[]'
            , '仅购买了仙魔礼包': '[]'
            , '购买了仙魔+常规礼包': '[]'
            , '仅购买了常规礼包': '[]'
            , '无任何共买记录': '[]'
        }

    # @BaseTransform.merge_cells_general
    
    @modify_cell_matched
    def prir_edit(self, *args, **kwargs):
        pass

    @df2ws
    def on_edit(self, *args, **kwargs):
        pass

    @merge_cells_general
    def post_edit(self, *args, **kwrags):
        pass

    def process_v1(self, ws, df, start_row, *args, **kwargs):
        if 'modify_dict' in kwargs:
            df = self.prir_edit(df, modify_dict = kwargs['modify_dict'])
        # ret_row = super(LevelFailureTransform, self).run(ws, df, start_row)
        ret_row = self.on_edit(ws, df, start_row)
        if 'groupby_target' in kwargs:
            self.post_edit(ws, df, start_row, groupby_target = kwargs['groupby_target'])
        else:
            self.post_edit(ws, df, start_row, groupby_target = self.groupby_target)
        return ret_row

    def run(self, ws, df, start_row):
        return self.process_v1(ws, df, start_row, modify_dict = self.modify_dict)

class SpiritFailureTransform(LevelFailureTransform):
    def __init__(self, *args, **kwargs):
        super(LevelFailureTransform, self).__init__()
        self.groupby_target = [
            (['日期', '神识等级'], '玩家名单')
        ]
    
    def process_v1(self, ws, df, start_row, *args, **kwargs):
        return super().process_v1(ws, df, start_row, groupby_target=self.groupby_target)

    def run(self, ws, df, start_row):
        return self.process_v1(ws, df, start_row)

class BreakthruLevelFailureIntervalConsumption:

    # add a descriptor that validate the config dict
    # the dict must contains df, ws, start_row as first 3 args

    def __init__(self, *args, **kwargs) -> None:
        self.groupby_target = [
            (['失败次数', '礼包去重序列'], '玩家名单'),
            (['失败次数', '玩家名单'], '礼包去重序列')
        ]
        self.args = {
            'outer_col': '日期',
            'inner_col': '等级名称',
            'drop_col': ['境界等级'],
            'drop_cols': ['等级名称', '日期'],
            'insert': '{outer_col}: {dt} {inner_col}: {level}',
            'groupby_target': [
                    (['失败次数', '礼包去重序列'], '玩家名单'),
                    (['失败次数', '玩家名单'], '礼包去重序列')
            ],
            'alignment': 'general,center',
        }

    @drop_columns
    def drop(self, *args, **kwargs):
        pass

    @select_columns
    def select(self, df, columns_dict, *args, **kwargs):
        pass

    @insert_row
    def insert(self, *args, **kwargs):
        pass

    @df2ws
    @merge_cells_general
    def edit_ws(self, ws, df, start_row, *args, **kwargs):
        print('enter edit_ws')
        pass
    
    def process_v1(self, ws, df, start_row):
        """
            step1: drop '境界等级'
            step2: iterate '日期' '等级名称'
                select cols
                drop '等级名称' '日期'
                insert line
                edit_ws: df2ws, merge_cells_general
        
        """
        start_row = start_row
        # start_row = self.insert(ws, start_row, f'下表为突破失败（后续可能成功或失败或当日结束）后仙魔大陆礼包商品兑换情况')
        # start_row = self.insert(ws, start_row, f'【礼包去重序列】（当日总失败次数值下的所有突破失败后兑换仙魔礼包去重），分组粒度为【失败次数】；【玩家名单】（当日总失败次数值下的所有突破失败的玩家名单汇总），分组粒度为【失败次数】')
        df = self.drop(df, ['境界等级'])
        for dt in df['日期'].unique():
            for level in df[df['日期']==dt]['等级名称'].unique():
                columns_dict = {
                    '日期': dt,
                    '等级名称': level
                }
                _df = self.select(df, columns_dict)
                _df = self.drop(_df, ['等级名称', '日期'])
                start_row = self.insert(ws, start_row, f'日期：{dt} 境界等级: {level}')
                ret_row = self.edit_ws(ws, _df, start_row, 
                                       groupby_target = self.groupby_target,
                                       alignment = 'general,center'
                )
                start_row = ret_row
        return ret_row
    
    def process_v3(self, ws, df, start_row, *args, **kwargs):
        start_row = start_row
        # start_row = self.insert(ws, start_row, f'下表为突破失败（后续可能成功或失败或当日结束）后仙魔大陆礼包商品兑换情况')
        # start_row = self.insert(ws, start_row, f'【礼包去重序列】（当日总失败次数值下的所有突破失败后兑换仙魔礼包去重），分组粒度为【失败次数】；【玩家名单】（当日总失败次数值下的所有突破失败的玩家名单汇总），分组粒度为【失败次数】')
        drop_col = self.args['drop_col']
        df = self.drop(df, drop_col)
        outer_col = self.args['outer_col']
        inner_col = self.args['inner_col']
        for dt in df[outer_col].unique():
            for level in df[df[outer_col]==dt][inner_col].unique():
                columns_dict = {
                    outer_col: dt,
                    inner_col: level
                }
                _df = self.select(df, columns_dict)

                drop_cols = self.args['drop_cols']
                _df = self.drop(_df, drop_cols)

                start_row = self.insert(ws, start_row, f'{outer_col}: {dt} {drop_col[0]}: {level}')

                groupby_target = self.args['groupby_target']
                alignment = self.args['alignment']
                ret_row = self.edit_ws(ws, _df, start_row, 
                                       groupby_target = groupby_target,
                                       alignment = alignment
                )
                start_row = ret_row
        return ret_row  

    def process_v2(self, ws, df, start_row, *args, **kwargs):
        """
            process_v1 generalized version, paramatize & decouple each steps v1 to optimize the function's scalability
            default params:
            ws: worksheet
            df: dataframe
            start_row: int

            actual steps:
            1. prior edit: drop columns
            2. on edit: select columns
            3. post edit: insert row
        """
        kwargs = {
            'outer_col': '日期',
            'inner_col': '等级名称',
            'prior_edit': {
                self.drop: ['境界等级']
            },
            'groupby_target': [
                    (['失败次数', '礼包去重序列'], '玩家名单'),
                    (['失败次数', '玩家名单'], '礼包去重序列')
            ],
            'alignment': 'general,center',
            'on_edit': {
                self.drop: ['等级名称', '日期'],
                self.insert: '日期：{i} 境界等级: {j}'

            }
        }

        if 'prior_edit' in kwargs:
            for func, val in kwargs['prior_edit'].items():
                df = func(df, val, *args, **kwargs)

        if 'outer_col' in kwargs and 'inner_col' in kwargs:
            outer_col = kwargs['outer_col']
            inner_col = kwargs['inner_col']
            for i in df[outer_col]:
                for j in df[inner_col]:
                    _columns_dict = {
                        outer_col: i,
                        inner_col: j
                    }
                    _df = self.select(df, _columns_dict)
                    
                    if 'on_edit' in kwargs:
                        for func, val in kwargs['on_edit'].items():
                            if type(val) == str:

                                val = val.format(i=i,j=j)
                                # insert方法调用，参数协议不对
                            _df = func(_df, val, *args, **kwargs)

                    ret_row = self.edit_ws(ws, _df, start_row, 
                        groupby_target = kwargs['groupby_target'],
                        alignment = kwargs['alignment']
                    )

                    if 'post_edit' in kwargs:
                        for func, col in kwargs['post_edit'].items():
                            _df = func(_df, col, *args, **kwargs)
        return ret_row

    def run(self, ws, df, start_row, *args, **kwargs):
        return self.process_v3(ws, df, start_row, *args, **kwargs)

class BreakthruSpiritFailureIntervalConsumption(BreakthruLevelFailureIntervalConsumption):
    def __init__(self, *args, **kwargs) -> None:
        self.args = {
            'outer_col': '日期',
            'inner_col': '等级名称',
            'drop_col': ['神识等级'],
            'drop_cols': ['等级名称', '日期'],
            'insert': '{outer_col}: {dt} {inner_col}: {level}',
            'groupby_target': [
                    (['失败次数', '礼包去重序列'], '玩家名单'),
                    (['失败次数', '玩家名单'], '礼包去重序列')
            ],
            'alignment': 'general,center',
        }

    def process_v1(self, ws, df, start_row, *args, **kwargs):
        return super().process_v3(ws, df, start_row, *args, **kwargs)
    
    def run(self, ws, df, start_row):
        return self.process_v1(ws, df, start_row)

class VipFisheryDistribution:
    def __init__(self, *args, **kwargs) -> None:
        """
            consider add type check descriptor to validate the sum_columns type
                check: type -> int
                check: not null
        """
        self.args = {
            'sum_columns': ['所有渔场都玩玩家数', '只常规渔场玩家数', '只玩仙魔玩家数']
            ,'sum_new_column_name': '合计（人数）'
            ,'percent_columns': ['所有渔场都玩占比', '只玩常规渔场占比', '只玩仙魔渔场占比']
            ,'percent_new_column_name': ['所有渔场都玩占比', '只玩常规渔场占比', '只玩仙魔渔场占比']
            ,'remaining_columns': ['日期', 'VIP']
            ,'modify_dict':{
                'VIP': 'str'
            }
        }
    
    from utils_dateframe import sum_columns, ratio_columns, drop_columns, columns_switch, modify_column_type

    @modify_column_type
    def modify_column_type(self, *args, **kwargs):
        print('modify_column_type')
        pass

    def add_sum_row(self, df):
        # 只支持单日dt
        sums = df[self.args['sum_columns']].sum()
        new_row = []
        for remain_col in self.args['remaining_columns']:
            if remain_col == '日期':
                new_row.append(df[remain_col].iloc[-1])
            if remain_col == 'VIP':
                new_row.append('总计')
        for col in self.args['sum_columns']:
            new_row.append(sums[col])
        df = pd.concat([df, pd.DataFrame([new_row], columns=df.columns)], ignore_index=True)
        return df

    @sum_columns
    def get_sum_columns(self, *args, **kwargs):
        pass
    
    @ratio_columns
    def get_ratio_columns(self, *args, **kwargs):
        pass

    @df2ws
    def on_edit(self, *args, **kwargs):
        pass

    @drop_columns
    def drop_columns(self, *args, **kwargs):
        pass

    @columns_switch
    def switch_columns(self, *args, **kwargs):
        pass

    def test_called_by_visualizer(self):
        print("test_called_by_visualizer")
        print(self.args)

    def process_df_alldata(self, df, *args, **kwargs):
        df = self.modify_column_type(df, self.args['modify_dict'])
        df = self.add_sum_row(df)
        df = self.get_sum_columns(
            df, 
            self.args['sum_columns'], 
            self.args['sum_new_column_name']
        )
        df = self.get_ratio_columns(
            df, 
            self.args['sum_columns'], 
            self.args['sum_new_column_name'],
            self.args['percent_columns']
        )
        return df

    def process_df(self, df, *args, **kwargs):
        df = self.modify_column_type(df, self.args['modify_dict'])
        df = self.add_sum_row(df)
        df = self.get_sum_columns(
            df, 
            self.args['sum_columns'], 
            self.args['sum_new_column_name']
        )
        df = self.get_ratio_columns(
            df, 
            self.args['sum_columns'], 
            self.args['sum_new_column_name'],
            self.args['percent_columns']
        )
        if 'all_data' in kwargs:
            return df
        df = self.drop_columns(
            df, 
            self.args['sum_columns']
        )
        df = self.switch_columns(
            df, 
            [self.args['sum_new_column_name']], 
            self.args['percent_columns']
        )
        return df

    def process_v1(self, ws, df, start_row, *args, **kwargs):
        df = self.process_df(df, *args, **kwargs)
        ret_row = self.on_edit(ws, df, start_row)
        # ret_row = self.on_edit(ws, df, start_row, self.args['sum_columns'], self.args['sum_new_column_name'])
        return ret_row

    def run(self, ws, df, start_row):
        return self.process_v1(ws, df, start_row)
    
class LevelSpiritPivotTransform:
    def __init__(self, *args, **kwargs) -> None:
        self.args = {
            'configfile': 'config_level_lv_spirit_facts.xlsx',
            'index': '境界等级',
            'column': '神识等级',
            'value': '用户id',
            'margin': '总计',
            'type1': 1,
            'type2': 2,
            'key': 'level',
            'name': 'level_name',
            'filter_key': 'isbreak',
            'isReplaceZero': {'all': 0},
            'style':{
                'row_set_fill': {
                    'color': 'FFFF0000',
                    'row_set': []
                },
                'bold_section': {
                    'start_row': -1,
                    'end_row': -1,
                    'column_index': -1
                }
            },
            'info': '注：仅统计同时具有境界等级和神识等级的用户'
        }
        # if kwargs is not empty
        if kwargs:
            print('created LevelSpiritPivotTransform with kwargs')
            print('kwargs: ', kwargs)
            _ = next(iter(kwargs.items()))
            import os
            print(_[1])
            config_file = os.path.join(_[1][-1] ,self.args.get('configfile'))
            self.end_dt = _[1][-2]
            
            self.args['config_file'] = config_file
            print('config_file: ', config_file)
        else:
            print('created LevelSpiritPivotTransform without kwargs')
            # raise FileNotFoundError(f'{self.args['configfile']} File not found')


    from utils_dateframe import modify_cell_matched_v2
    from utils_worksheet import color_cells, border_cell_column

    @modify_cell_matched_v2
    def prir_edit_df(self, *args, **kwargs):
        pass

    def process_df(self, df, df_config, *args, **kwargs):
        """
            pivot table
            file: utils_dataframe -> func pivot table
            file: tils_dataframe -> func sort columns, sort rows, join name
            args: feature1, feature2, value, isSum, SumName
            *args: sort
        """
        print(f'df: first 10: {df.head(10)}')
        config_type1 = df_config[df_config['type'] == self.args['type1']]
        print(config_type1)
        print(config_type1[self.args['key']].tolist()+ [self.args['margin']])
        # print(f'config_type1: first 10: {config_type1.head(10)}')
        
        config_type2 = df_config[df_config['type'] == self.args['type2']]

        # print(f'config_type2: first 10: {config_type2.head(10)}')

        pivot_table = df.pivot_table(
                        index=self.args['index'], 
                        columns=self.args['column'], 
                        values=self.args['value'], 
                        aggfunc=pd.Series.nunique, 
                        fill_value=0, 
                        margins=True, 
                        margins_name=self.args['margin']
                    )

        new_index = config_type1[self.args['key']].tolist() + [self.args['margin']]
        new_column = config_type2[self.args['key']].tolist() + [self.args['margin']]
        pivot_table = pivot_table.reindex(index=new_index, 
                                          columns=new_column, 
                                          fill_value=0)
        pivot_table = pivot_table.reset_index()

        pivot_table[self.args['index']] = pivot_table[self.args['index']] \
            .replace(config_type1.set_index(self.args['key'])[self.args['name']])
        columns_mapping = {level: name for level, name in zip(config_type2[self.args['key']], config_type2[self.args['name']])}
        pivot_table = pivot_table.rename(columns=columns_mapping)
        pivot_table = pivot_table.rename(columns={self.args['index']: '等级'})

        pivot_table = self.prir_edit_df(pivot_table, **self.args['isReplaceZero'])

        # self.add_sum_row(pivot_table)
        # print(f'pivot table: first 10: {pivot_table.head(10)}')
        
        return pivot_table

    @df2ws
    def prior_edit_ws(self, *args, **kwargs):
        pass

    @color_cells
    def on_edit_ws(self, *args, **kwargs):
        pass
    
    @insert_row
    def insert_row(self, *args, **kwargs):
        pass

    @border_cell_column
    def border_cell_column(self, *args, **kwargs):
        if not args:
            raise IndexError('func: border_cell_column: args is empty')
        section_dict = args[0]
        return section_dict['start_row'], section_dict['end_row'], section_dict['column_index']

    
    def process_ws(self, ws, df, start_row, df_config, *args, **kwargs):
        """
            ws generate and style
        """
        ret_row = self.insert_row(ws, start_row, f'日期: {self.end_dt}')

        start_row = ret_row
        ret_row = self.prior_edit_ws(ws, df, start_row, **{'regular': True})

        self.args['style']['bold_section'].update({'start_row': start_row, 'end_row': ret_row, 'column_index': 1})

        self.border_cell_column(ws, start_row, **self.args['style'])
        # ret_row = self.on_edit_ws(ws, df, ret_row, **self.args['style'])
        # ret_row = self.insert_row(ws, ret_row, self.args['info'])
        ret_row = self.insert_row(ws, ret_row, self.args['info'])
        # initialize the row set
        # df_break = df_config[df_config['isbreak'] == 1][[]]
        # df.merge(df_break, on)
        # row_set = [4, 5, 8, 9]
        # self.args['style']['row_set_fill'].update({'row_set': row_set})
        # self.on_edit_ws(ws, df, start_row, **self.args['style'])

        return ret_row

    def process_v1(self, ws, df, start_row, *args, **kwargs):
        """
            officially use for v1
        """
        df_config = pd.read_excel(self.args['config_file'])

        df = self.process_df(df, df_config)
        ret_row = self.process_ws(ws, df, start_row, df_config)

        return ret_row

    def run(self, ws, df, start_row):
        return self.process_v1(ws, df, start_row)
